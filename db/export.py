import io
from aiogram import Router
from aiogram.filters import Command
from aiogram.types import Message
from aiogram.types.input_file import BufferedInputFile
from openpyxl import Workbook

from db.database import fetch_all
from services.admin import suppress_group

export_router = Router()


def safe_sheet_title(title: str) -> str:
    forbidden = ['/', '\\', '?', '*', '[', ']']
    for ch in forbidden:
        title = title.replace(ch, '_')
    return title[:31]

# учзщке дщфвы
@export_router.message(Command("export_excel"))
@suppress_group
async def export_excel(message: Message):
    wb = Workbook()
    ws = wb.active
    ws.title = "Loads"

    ws.append(["Driver", "Dispatcher", "broker", "Load Number", "Rate", "Miles"])

    rows = await fetch_all("""
        SELECT d.name AS driver, ds.name AS dispatcher, l.broker, l.load_number, l.rate, l.miles
        FROM loads l
        JOIN drivers d ON l.driver_id = d.id
        JOIN dispatchers ds ON l.dispatcher_id = ds.id
    """)

    for row in rows:
        ws.append([row['driver'], row['dispatcher'], row['broker'], row['load_number'], row['rate'], row['miles']])

    buffer = io.BytesIO()
    wb.save(buffer)

    file = BufferedInputFile(
        buffer.getvalue(),
        filename="loads.xlsx"
    )

    await message.reply_document(
        document=file,
        caption="📊 export loads"
    )


# export by dispatch

@export_router.message(Command("export_dispatcher"))
@suppress_group
async def export_dispatcher(message: Message):
    text = message.text or ""
    dispatcher_name = text.replace("/export_dispatcher", "", 1).strip()

    if not dispatcher_name:
        await message.reply("❌ Please provide the dispatcher's name:\n/export_dispatcher Sam Walter")
        return

    rows = await fetch_all(
        """
            SELECT 
                d.name AS driver,
                ds.name AS dispatcher,
                l.load_number,
                l.rate,
                l.miles
            FROM loads l
            JOIN drivers d ON l.driver_id = d.id
            JOIN dispatchers ds ON l.dispatcher_id = ds.id
            WHERE ds.name = $1
        """,
        dispatcher_name,
    )

    if not rows:
        await message.reply(f"❌ Loads for dispather *{dispatcher_name}* not found", parse_mode="Markdown")
        return

    wb = Workbook()
    ws = wb.active
    sheet_name = safe_sheet_title(dispatcher_name)
    ws.title = sheet_name

    ws.append(["Driver", "Dispatcher", "Load Number", "Rate", "Miles"])

    total = 0
    total_miles = 0
    for row in rows:
        ws.append([row["driver"], row["dispatcher"], row["load_number"], row["rate"], row["miles"]])
        total += row["rate"] or 0
        total_miles += row["miles"] or 0

    # Итоговая строка
    ws.append([])
    ws.append(["", "", "TOTAL", round(total, 2), round(total_miles, 2)])

    buffer = io.BytesIO()
    wb.save(buffer)

    file = BufferedInputFile(
        buffer.getvalue(),
        filename=f"dispatcher_{dispatcher_name}.xlsx"
    )

    await message.reply_document(
        document=file,
        caption=f"📊 Gross dispatch: *{dispatcher_name}*\n💰 Total: ${total:.2f}\nMiles: {total_miles:.2f}",
        parse_mode="Markdown"
    )

# driver
@export_router.message(Command("export_driver"))
@suppress_group
async def export_driver(message: Message):
    text = message.text or ""
    driver_name = text.replace("/export_driver", "", 1).strip()

    if not driver_name:
        await message.reply("❌ Enter the driver's group name:\n/export_driver 015 Daniiar Zhunusov")
        return

    rows = await fetch_all(
        """
            SELECT
                d.name AS driver,
                ds.name AS dispatcher,
                l.load_number,
                l.rate,
                l.miles
            FROM loads l
            JOIN drivers d ON l.driver_id = d.id
            JOIN dispatchers ds ON l.dispatcher_id = ds.id
            WHERE d.name = $1
        """,
        driver_name,
    )

    if not rows:
        await message.reply(
            f"❌ Loads for driver *{driver_name}* not found",
            parse_mode="Markdown"
        )
        return

    wb = Workbook()
    ws = wb.active
    sheet_name = safe_sheet_title(driver_name)
    ws.title = sheet_name

    ws.append(["Driver", "Dispatcher", "Load Number", "Rate", "Miles"])

    total = 0
    total_miles = 0
    for row in rows:
        ws.append([row["driver"], row["dispatcher"], row["load_number"], row["rate"], row["miles"]])
        total += row["rate"] or 0
        total_miles += row["miles"] or 0

    ws.append([])
    ws.append(["", "", "TOTAL", round(total, 2), round(total_miles, 2)])

    buffer = io.BytesIO()
    wb.save(buffer)

    file = BufferedInputFile(
        buffer.getvalue(),
        filename=f"driver_{driver_name}.xlsx"
    )

    await message.reply_document(
        document=file,
        caption=f"🚚 Gross driver *{driver_name}*\n💰 Total: ${total:.2f}\nMiles: {total_miles:.2f}",
        parse_mode="Markdown"
    )
